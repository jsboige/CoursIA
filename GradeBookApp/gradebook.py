import pandas as pd
import numpy as np
import os
import glob
import re
from unidecode import unidecode
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from rapidfuzz import fuzz

# --- Paramètres globaux ---
number_of_projects = 2
nb_eval_fields_per_project = [3, 3]
professor_email = "jsboige@gmail.com"
PROFESSOR_FULL_NAME = "Jean-Sylvain Boige"
SIMILARITY_THRESHOLD = 80 # Seuil pour le fuzzy matching
TEACHER_WEIGHT = 1.0 # Poids de la note du professeur. 1.0 = 50% du total.

# --- Structures de données ---

class StudentRecord:
    def __init__(self, prenom, nom, sujet_projet_1, sujet_projet_2=None):
        self.prenom = prenom
        self.nom = nom
        self.sujets = [sujet_projet_1]
        if sujet_projet_2:
            self.sujets.append(sujet_projet_2)
        self.notes = []

    @property
    def moyenne(self):
        return np.mean(self.notes) if self.notes else 0

class EvaluationRecord:
    def __init__(self, date, email, nom, prenom, groupe, sujet_libre, notes_dict,
                 points_positifs, points_negatifs, recommandations):
        self.date = pd.to_datetime(date, dayfirst=True, errors='coerce')
        self.email = email
        self.nom = nom
        self.prenom = prenom
        self.groupe = str(groupe) if pd.notna(groupe) else ""
        self.sujet_libre = sujet_libre
        # Utiliser float() au lieu de int() pour préserver les décimales (ex: 9.5)
        self.notes = {k: float(v) for k, v in notes_dict.items() if pd.notna(v) and str(v).strip()}
        self.points_positifs = points_positifs
        self.points_negatifs = points_negatifs
        self.recommandations = recommandations
        self.is_teacher = False

    @property
    def note(self):
        if not self.notes:
            return 0
        total_notes = sum(self.notes.values())
        nb_eval_fields = len(self.notes)
        # Each criterion is out of 10. We normalize to a score out of 20.
        # (total_score / (num_fields * 10)) * 20  which simplifies to (total_score * 2) / num_fields
        return (total_notes * 2) / nb_eval_fields if nb_eval_fields > 0 else 0

class GroupEvaluation:
    def __init__(self, groupe, evaluations, group_members):
        self.groupe = groupe
        self.evaluations = evaluations
        self.group_members = group_members
        self.note_rectifiee = 0

    @property
    def moyenne(self):
        student_evals = [e.note for e in self.evaluations if not e.is_teacher]
        teacher_evals = [e.note for e in self.evaluations if e.is_teacher]

        student_avg = np.mean(student_evals) if student_evals else 0
        teacher_avg = np.mean(teacher_evals) if teacher_evals else 0
        
        # Si pas de notes étudiantes, on retourne la note du prof directement
        if not student_evals:
            return teacher_avg

        # Si pas de note prof, on retourne la moyenne étudiante
        if not teacher_evals:
            return student_avg
            
        # Logique de pondération du notebook original
        # Note: avec TEACHER_WEIGHT = 1.0, la note du prof pèse 50% de la note finale
        # peu importe le nombre d'évaluations étudiantes.
        numerator = student_avg + (teacher_avg * TEACHER_WEIGHT)
        denominator = 1 + TEACHER_WEIGHT
        
        return numerator / denominator if denominator > 0 else 0

    @property
    def ecart_type(self):
        all_notes = [e.note for e in self.evaluations]
        if len(all_notes) < 2:
            return 0
        return np.std(all_notes)

    @property
    def date(self):
        return self.evaluations[len(self.evaluations) // 2].date if self.evaluations else None


class ProjectEvaluation:
    def __init__(self, professor_email, student_records):
        self.professor_email = professor_email
        self.student_records = student_records
        self.grouped_evaluations = []

    @staticmethod
    def fold_accents(text):
        return unidecode(text)

    def project_name(self, name):
        if not isinstance(name, str):
            return ""
        first_name = name.split(" ")[0].split("-")[0]
        first_name = first_name[:5]
        return self.fold_accents(first_name.lower().strip())

    def match_names(self, student_or_eval1, eval2):
        name1 = student_or_eval1.nom
        prenom1 = student_or_eval1.prenom
        
        name2 = eval2.nom
        prenom2 = eval2.prenom

        return (self.project_name(name1) == self.project_name(name2) and self.project_name(prenom1) == self.project_name(prenom2)) or \
               (self.project_name(prenom1) == self.project_name(name2) and self.project_name(name1) == self.project_name(prenom2))

    def is_valid_evaluation(self, group_evaluation, eval_record, seen_evaluators):
        # Règle 1: Rejeter les notes invalides
        # Le professeur peut mettre 20, mais pas les étudiants (limités à 19.999...)
        is_student_perfect_score = not eval_record.is_teacher and eval_record.note >= 20
        if is_student_perfect_score or eval_record.note < 1:
            print(f"AVERTISSEMENT: Note invalide ({eval_record.note}) pour le groupe {group_evaluation.groupe} par {eval_record.prenom} {eval_record.nom}. Evaluation écartée.")
            return False

        # Règle 2: Rejeter les dates incohérentes
        if group_evaluation.date and abs((group_evaluation.date - eval_record.date).total_seconds()) > 5 * 3600:
            print(f"AVERTISSEMENT: Date d'évaluation incohérente pour le groupe {group_evaluation.groupe} par {eval_record.prenom} {eval_record.nom}. Evaluation écartée.")
            return False

        # Règle 3: Rejeter les évaluateurs non-inscrits (sauf le professeur)
        if not eval_record.is_teacher:
            is_registered = any(self.match_names(student, eval_record) for student in self.student_records)
            if not is_registered:
                print(f"AVERTISSEMENT: L'évaluateur {eval_record.prenom} {eval_record.nom} n'est pas inscrit au cours. Evaluation pour {group_evaluation.groupe} écartée.")
                return False

        # Règle 4: Rejeter les auto-évaluations
        is_member = any(self.match_names(member, eval_record) for member in group_evaluation.group_members)
        if is_member:
            print(f"AVERTISSEMENT: Auto-évaluation par {eval_record.prenom} {eval_record.nom} pour son propre groupe {group_evaluation.groupe}. Evaluation écartée.")
            return False

        # Règle 5: Rejeter les évaluations en double
        evaluator_id = eval_record.email
        if evaluator_id in seen_evaluators:
            print(f"AVERTISSEMENT: Evaluation double par {eval_record.prenom} {eval_record.nom} pour le groupe {group_evaluation.groupe}. Evaluation écartée.")
            return False

        return True

    def filter_evaluations(self):
        for group_eval in self.grouped_evaluations:
            valid_evaluations = []
            seen_evaluators = set()
            # Trier les évaluations par date pour que la première évaluation valide serve de référence temporelle
            sorted_evaluations = sorted(group_eval.evaluations, key=lambda e: e.date)
            for evaluation in sorted_evaluations:
                if self.is_valid_evaluation(group_eval, evaluation, seen_evaluators):
                    valid_evaluations.append(evaluation)
                    seen_evaluators.add(evaluation.email)
            group_eval.evaluations = valid_evaluations

    @property
    def moyenne(self):
        if not self.grouped_evaluations: return 0
        notes = [ge.moyenne * len(ge.group_members) for ge in self.grouped_evaluations if ge.group_members]
        counts = [len(ge.group_members) for ge in self.grouped_evaluations if ge.group_members]
        return np.sum(notes) / np.sum(counts) if np.sum(counts) > 0 else 0

    @property
    def ecart_type(self):
        if not self.grouped_evaluations: return 0
        mean = self.moyenne
        squared_diffs = [((ge.moyenne - mean) ** 2) * len(ge.group_members) for ge in self.grouped_evaluations if ge.group_members]
        counts = [len(ge.group_members) for ge in self.grouped_evaluations if ge.group_members]
        return np.sqrt(np.sum(squared_diffs) / np.sum(counts)) if np.sum(counts) > 0 else 0

# --- Fonctions de traitement ---

def load_student_records(file_path, column_mapping=None):
    """
    Charge les étudiants depuis un fichier CSV avec mapping de colonnes personnalisable.
    
    Args:
        file_path: Chemin du fichier CSV d'inscription
        column_mapping: Dictionnaire optionnel de mapping des colonnes.
                       Si None, utilise le mapping par défaut EPITA.
    
    Returns:
        Liste d'objets StudentRecord
    """
    df = pd.read_csv(file_path, delimiter=',')
    
    # Mapping par défaut (EPITA) si non fourni
    if column_mapping is None:
        column_mapping = {
            # Variations pour le prénom
            "Prénom": "prenom",
            "Prenom": "prenom",
            # Variations pour le nom
            "Nom": "nom",
            # Variations pour le sujet du projet
            "Sujet IA Symbolique": "sujet_projet_1",
            "Sujet Projet de Programmation par Contrainte": "sujet_projet_1"
        }
    
    df = df.rename(columns=column_mapping)
    
    students = []
    # Vérifier que les colonnes nécessaires existent après renommage
    required_cols = ['prenom', 'nom', 'sujet_projet_1']
    if not all(col in df.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in df.columns]
        raise ValueError(f"Colonnes manquantes dans {file_path} après renommage : {missing_cols}. Colonnes disponibles : {df.columns.tolist()}")

    for _, row in df.iterrows():
        # Vérification de la validité des données avant de créer l'objet
        prenom = row.get('prenom')
        nom = row.get('nom')
        projet = row.get('sujet_projet_1')
        
        if pd.isna(prenom) or not str(prenom).strip() or pd.isna(nom) or not str(nom).strip():
            print(f"AVERTISSEMENT: Ligne ignorée dans {file_path} car le nom ou le prénom est manquant : Ligne { _ + 2 }")
            continue
        
        if pd.isna(projet) or not str(projet).strip():
            continue
            
        students.append(StudentRecord(prenom, nom, projet))
    return students


def load_grades_from_file(file_path, correct_prof_name, correct_prof_email, threshold=80):
    """
    Charge les évaluations depuis un fichier CSV, identifie le professeur avec une logique de fuzzy matching
    et retourne une liste d'objets EvaluationRecord enrichis.
    """
    df = pd.read_csv(file_path, delimiter=',')
    column_mapping = {
        "Horodateur": "date", "Adresse e-mail": "email", "Votre nom": "nom", "Votre prénom": "prenom",
        "Groupe à évaluer": "groupe", "Sujet de la présentation": "sujet_libre",
        "Qualité de la présentation (communication, la forme)": "note_communication",
        "Qualité théorique (principes utilisés, classe d'algorithmes, contexte et explications des performances et des problèmes, histoire etc.)": "note_theorique",
        "Qualité technique (livrables, commits, qualité du code, démos, résultats, perspectives)": "note_technique",
        "Organisation (planning, répartition des tâches, collaboration, intégration au projet Github)": "note_organisation",
        "Points positifs de la présentation": "points_positifs",
        "Points négatifs de la présentation": "points_negatifs",
        "Recommandations pour s'améliorer": "recommandations"
    }
    df.rename(columns=column_mapping, inplace=True)

    df['nom_complet'] = df['prenom'].astype(str).str.strip() + " " + df['nom'].astype(str).str.strip()

    def normalize_name(name):
        return unidecode(name).lower().strip()

    normalized_prof_name = normalize_name(correct_prof_name)

    evaluations = []
    note_cols = [col for col in ["note_communication", "note_theorique", "note_technique", "note_organisation"] if col in df.columns]

    for _, row in df.iterrows():
        if not any(pd.notna(row.get(col)) and str(row.get(col)).strip() for col in note_cols):
            continue

        evaluator_name = row['nom_complet']
        normalized_evaluator = normalize_name(evaluator_name)
        similarity = fuzz.ratio(normalized_prof_name, normalized_evaluator)
        
        is_teacher = False
        if similarity >= threshold:
            is_teacher = True
            if row['email'] and row['email'].strip() != correct_prof_email:
                print(f"AVERTISSEMENT : L'évaluateur '{evaluator_name}' identifié comme le professeur a une adresse email potentiellement incorrecte: '{row['email']}'")

        notes_dict = {col: row.get(col) for col in note_cols}
        record = EvaluationRecord(
            row.get('date'), row.get('email'), row.get('nom'), row.get('prenom'),
            row.get('groupe'), row.get('sujet_libre'), notes_dict,
            row.get('points_positifs'), row.get('points_negatifs'), row.get('recommandations')
        )
        record.is_teacher = is_teacher
        evaluations.append(record)

    return evaluations


def adjust_grade(original_grade, group_mean, group_std_dev, target_mean=10, target_std_dev=2):
    """Fonction de centrage-réduction pour normaliser les notes."""
    if group_std_dev == 0:
        return target_mean
    adjusted_grade = ((original_grade - group_mean) / group_std_dev) * target_std_dev + target_mean
    return np.clip(adjusted_grade, 0, 20)


group_size_adjustments = {1: 3.0, 2: 1.0, 3: 0.0, 4: -1.0}

def palier_group_size_adjustment(group_size):
    """Retourne le bonus/malus selon la taille du groupe."""
    # Malus par défaut de -2 pour les tailles de groupe non définies (ex: 5, 6...)
    return group_size_adjustments.get(group_size, -2.0)


def fuzzy_match_group(eval_group_name, student_project_string):
    """
    Tente de faire correspondre un nom de groupe d'évaluation (potentiellement simple)
    avec une chaîne de projet d'étudiant (potentiellement plus descriptive).
    """
    if not isinstance(eval_group_name, str) or not isinstance(student_project_string, str):
        return False

    # Normalisation : minuscules, suppression des espaces superflus
    eval_norm = eval_group_name.lower().strip()
    student_norm = student_project_string.lower().strip()

    # Cas 1: Correspondance directe (pour les noms simples comme "Quoridor")
    if eval_norm == student_norm:
        return True

    # Cas 2: Si l'un est contenu dans l'autre (ex: "groupe 1" et "1")
    if eval_norm in student_norm or student_norm in eval_norm:
        return True

    # Cas 3: Extraction intelligente des numéros de groupe
    # On cherche le pattern "groupe X" ou "group X" pour extraire le numéro principal
    group_pattern = r'groupe?\s*(\d+)'
    
    eval_group_match = re.search(group_pattern, eval_norm)
    student_group_match = re.search(group_pattern, student_norm)
    
    # Si les deux ont un numéro de groupe explicite, on compare ces numéros
    if eval_group_match and student_group_match:
        return eval_group_match.group(1) == student_group_match.group(1)
    
    # Cas 4: Fallback - correspondance par contenu significatif (pour les noms de projet)
    # Extraire les mots significatifs (> 3 caractères)
    eval_words = set(word for word in re.findall(r'\b\w+\b', eval_norm) if len(word) > 3)
    student_words = set(word for word in re.findall(r'\b\w+\b', student_norm) if len(word) > 3)
    
    # Exclure les mots communs non significatifs
    common_words = {'groupe', 'group', 'projet', 'project', 'pour', 'avec', 'dans', 'pour', 'machine', 'learning'}
    eval_words -= common_words
    student_words -= common_words
    
    # S'il y a au moins 2 mots en commun, c'est probablement le même projet
    if len(eval_words.intersection(student_words)) >= 2:
        return True

    return False


def is_feedback_empty(evaluation):
    """
    Vérifie si un feedback est vide (tous les champs sont NaN, 'nan', ou vides).
    
    Args:
        evaluation: Objet EvaluationRecord
    
    Returns:
        True si le feedback est vide, False sinon
    """
    fields = ['points_positifs', 'points_negatifs', 'recommandations']
    
    for field in fields:
        value = getattr(evaluation, field, None)
        
        # Si la valeur est None ou NaN pandas
        if value is None or pd.isna(value):
            continue
        
        # Convertir en string et nettoyer
        str_value = str(value).strip().lower()
        
        # Si c'est une vraie valeur (pas 'nan', pas vide)
        if str_value and str_value != 'nan':
            return False
    
    # Tous les champs sont vides
    return True


def apply_rectification(project_eval, target_mean=15.0, target_stdev=2.0, skip_normalization=False):
    """
    Applique le redressement en 2 étapes (conforme au notebook GradeBook.ipynb) :
    1. Étape A : Bonus/malus selon taille du groupe
    2. Étape B : Centrage-réduction avec la moyenne et écart-type cibles (optionnel)

    Args:
        project_eval: Objet ProjectEvaluation
        target_mean: Moyenne cible pour le redressement
        target_stdev: Écart-type cible pour le redressement
        skip_normalization: Si True, saute l'étape B (centrage-réduction)
    """
    
    print("\n" + "="*80)
    print("ÉTAPE A : Application des bonus/malus selon la taille des groupes")
    print("="*80)
    
    # Étape A : Application du bonus/malus
    for group_eval in project_eval.grouped_evaluations:
        raw_avg = group_eval.moyenne  # Note brute pondérée
        group_size = len(group_eval.group_members)
        bonus_malus = palier_group_size_adjustment(group_size)
        
        adjusted = raw_avg + bonus_malus
        adjusted = np.clip(adjusted, 0, 20)
        
        group_eval.note_rectifiee = adjusted
        
        print(f"  {group_eval.groupe[:40]:40} | Taille: {group_size} | Brute: {raw_avg:.2f} | Bonus/Malus: {bonus_malus:+.1f} | Ajustée: {adjusted:.2f}")
    
    # Étape B : Centrage-réduction (optionnel)
    if skip_normalization:
        print("\n" + "="*80)
        print("ÉTAPE B : Centrage-réduction DÉSACTIVÉE (skip_normalization=True)")
        print("="*80)
        print("  Les notes conservent les valeurs après bonus/malus uniquement.")

        # Statistiques finales (sans normalisation)
        final_grades = [g.note_rectifiee for g in project_eval.grouped_evaluations]
        if final_grades:
            print(f"\n  Statistiques finales :")
            print(f"    Moyenne: {np.mean(final_grades):.2f}")
            print(f"    Écart-type: {np.std(final_grades):.2f}")
            print(f"    Min: {np.min(final_grades):.2f}")
            print(f"    Max: {np.max(final_grades):.2f}")
    else:
        print("\n" + "="*80)
        print("ÉTAPE B : Centrage-réduction statistique")
        print(f"  Moyenne cible: {target_mean:.2f} | Écart-type cible: {target_stdev:.2f}")
        print("="*80)

        # Calcul de la moyenne et écart-type du projet (basés sur les notes de l'étape A)
        total_students = sum(len(g.group_members) for g in project_eval.grouped_evaluations)

        if total_students == 0:
            print("⚠️  Aucun étudiant trouvé, impossible d'appliquer le redressement")
            return

        sum_rectified = sum(g.note_rectifiee * len(g.group_members) for g in project_eval.grouped_evaluations)
        project_mean = sum_rectified / total_students

        sum_squared_diffs = sum(
            ((g.note_rectifiee - project_mean) ** 2) * len(g.group_members)
            for g in project_eval.grouped_evaluations
        )
        project_stdev = np.sqrt(sum_squared_diffs / total_students)

        print(f"\n  Statistiques avant redressement :")
        print(f"    Moyenne actuelle: {project_mean:.2f}")
        print(f"    Écart-type actuel: {project_stdev:.2f}")
        print()

        # Application de la fonction AdjustGrade pour chaque groupe
        for group_eval in project_eval.grouped_evaluations:
            current_grade = group_eval.note_rectifiee
            final_grade = adjust_grade(current_grade, project_mean, project_stdev, target_mean, target_stdev)
            group_eval.note_rectifiee = final_grade

            print(f"  {group_eval.groupe[:40]:40} | Avant: {current_grade:.2f} | Après: {final_grade:.2f}")

        # Statistiques finales
        final_grades = [g.note_rectifiee for g in project_eval.grouped_evaluations]
        final_mean = np.mean(final_grades)
        final_stdev = np.std(final_grades)

        print(f"\n  Statistiques après redressement :")
        print(f"    Moyenne finale: {final_mean:.2f}")
        print(f"    Écart-type final: {final_stdev:.2f}")
        print(f"    Min: {np.min(final_grades):.2f}")
        print(f"    Max: {np.max(final_grades):.2f}")


def generate_excel_workbook(project_eval, student_records, output_path):
    """
    Génère le fichier Excel au format EPF avec 2 feuilles :
    - Feuille "Résumé Étudiants" : Nom, Prénom, Groupe Projet, Note Projet, Moyenne Finale
    - Feuille "Projet 1 Feedback" : Feedbacks filtrés (sans lignes NaN)
    
    Args:
        project_eval: Objet ProjectEvaluation
        student_records: Liste des StudentRecord
        output_path: Chemin de sortie du fichier Excel
    """
    
    print("\n" + "="*80)
    print("GÉNÉRATION DU FICHIER EXCEL")
    print("="*80)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut
    
    # === FEUILLE 1 : Résumé Étudiants ===
    summary_sheet = wb.create_sheet("Résumé Étudiants")
    
    # En-têtes
    summary_sheet.cell(1, 1, "Nom")
    summary_sheet.cell(1, 2, "Prénom")
    summary_sheet.cell(1, 3, "Groupe Projet")
    summary_sheet.cell(1, 4, "Note Projet")
    summary_sheet.cell(1, 5, "Moyenne Finale")
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, 6):
        cell = summary_sheet.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Trier les étudiants par nom et prénom
    sorted_students = sorted(student_records, key=lambda s: (s.nom, s.prenom))
    
    row = 2
    for student in sorted_students:
        # Trouver le groupe et la note de l'étudiant
        group_eval = None
        for g in project_eval.grouped_evaluations:
            if any(s.nom == student.nom and s.prenom == student.prenom for s in g.group_members):
                group_eval = g
                break
        
        if group_eval:
            summary_sheet.cell(row, 1, student.nom)
            summary_sheet.cell(row, 2, student.prenom)
            summary_sheet.cell(row, 3, group_eval.groupe)
            summary_sheet.cell(row, 4, round(group_eval.note_rectifiee, 1))
            summary_sheet.cell(row, 5, round(group_eval.note_rectifiee, 1))  # Ici = note projet car 1 seul projet
            
            # Format numérique
            summary_sheet.cell(row, 4).number_format = '0.0'
            summary_sheet.cell(row, 5).number_format = '0.0'
            
            row += 1
    
    # Ajuster les largeurs de colonnes
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 35
    summary_sheet.column_dimensions['D'].width = 12
    summary_sheet.column_dimensions['E'].width = 18
    
    print(f"✅ Feuille 'Résumé Étudiants' créée ({row - 2} étudiants)")
    
    # === FEUILLE 2 : Feedback du Projet (avec filtrage NaN) ===
    feedback_sheet = wb.create_sheet("Projet 1 Feedback")
    
    # En-têtes
    feedback_sheet.cell(1, 1, "Groupe à évaluer")
    feedback_sheet.cell(1, 2, "Points positifs")
    feedback_sheet.cell(1, 3, "Points négatifs")
    feedback_sheet.cell(1, 4, "Recommandations")
    
    # Style des en-têtes
    for col in range(1, 5):
        cell = feedback_sheet.cell(1, col)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Contenu des feedbacks avec filtrage NaN
    feedback_row = 2
    feedback_count = 0
    filtered_count = 0
    
    for group_eval in project_eval.grouped_evaluations:
        for evaluation in group_eval.evaluations:
            # FILTRAGE NaN : Ignorer les feedbacks complètement vides
            if is_feedback_empty(evaluation):
                filtered_count += 1
                continue
            
            feedback_sheet.cell(feedback_row, 1, group_eval.groupe)
            feedback_sheet.cell(feedback_row, 2, str(getattr(evaluation, 'points_positifs', '')))
            feedback_sheet.cell(feedback_row, 3, str(getattr(evaluation, 'points_negatifs', '')))
            feedback_sheet.cell(feedback_row, 4, str(getattr(evaluation, 'recommandations', '')))
            
            # Retour à la ligne automatique
            for col in range(1, 5):
                cell = feedback_sheet.cell(feedback_row, col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            feedback_row += 1
            feedback_count += 1
    
    # Ajuster les largeurs de colonnes
    feedback_sheet.column_dimensions['A'].width = 35
    feedback_sheet.column_dimensions['B'].width = 40
    feedback_sheet.column_dimensions['C'].width = 40
    feedback_sheet.column_dimensions['D'].width = 40
    
    print(f"✅ Feuille 'Projet 1 Feedback' créée ({feedback_count} feedbacks non-vides, {filtered_count} lignes NaN filtrées)")
    
    # Sauvegarder le fichier
    wb.save(output_path)
    print(f"\n✅ Fichier Excel généré : {output_path}")
    print(f"   - Feuille 'Résumé Étudiants' : {row - 2} étudiants")
    print(f"   - Feuille 'Projet 1 Feedback' : {feedback_count} feedbacks (après filtrage NaN)")


def generate_workbook(project_evaluations, students, output_path):
    """
    Génère un classeur Excel contenant le rapport de notes détaillé (format EPITA).
    """
    if not project_evaluations:
        print(f"ERREUR: Aucune donnée à traiter pour le rapport : {output_path}")
        return

    # On suppose qu'on traite un projet à la fois
    project_evaluation = project_evaluations[0]

    report_data = []
    for group_eval in project_evaluation.grouped_evaluations:
        members = ", ".join([f"{s.prenom} {s.nom}" for s in group_eval.group_members])
        report_data.append({
            "Groupe (Sujet)": group_eval.groupe,
            "Membres": members,
            "Taille du Groupe": len(group_eval.group_members),
            "Moyenne Brute": group_eval.moyenne,
            "Note Intermédiaire (avec bonus/malus)": getattr(group_eval, 'intermediate_grade', 'N/A'),
            "Note Finale Ajustée": group_eval.note_rectifiee
        })

    if not report_data:
        print(f"AVERTISSEMENT: Aucune donnée de groupe à écrire dans le rapport {output_path}.")
        # Créer un DataFrame vide avec les en-têtes pour éviter l'échec de la vérification
        df = pd.DataFrame(columns=["Groupe (Sujet)", "Membres", "Taille du Groupe", "Moyenne Brute", "Note Intermédiaire (avec bonus/malus)", "Note Finale Ajustée"])
    else:
        df = pd.DataFrame(report_data)

    try:
        # Utilisation de pandas pour écrire facilement le fichier Excel
        df.to_excel(output_path, sheet_name='Rapport de Notes', index=False, engine='openpyxl')
        print(f"Classeur généré avec succès à : {output_path}")
    except Exception as e:
        print(f"ERREUR CRITIQUE: Impossible de sauvegarder le fichier Excel à {output_path}. Exception: {e}")


def run_pipeline(config):
    """
    Fonction principale qui exécute le pipeline complet de notation (modèle EPF).
    
    Args:
        config: Dictionnaire de configuration contenant :
            - nom_classe: Nom de la classe (pour affichage)
            - inscriptions_path: Chemin du fichier CSV d'inscription
            - evaluations_path: Chemin du fichier CSV d'évaluations
            - output_path: Chemin du fichier Excel de sortie
            - column_mapping: Mapping des colonnes du CSV d'inscription
            - target_mean: Moyenne cible pour le redressement
            - target_std: Écart-type cible pour le redressement
            - teacher_weight: Poids de la note professeur (1.0 = 50%)
            - professor_email: Email du professeur
            - blacklisted_groups: Liste optionnelle de groupes à ignorer
    
    Returns:
        Objet ProjectEvaluation avec les résultats du traitement
    """
    
    # Extraction de la configuration
    nom_classe = config['nom_classe']
    inscriptions_path = config['inscriptions_path']
    evaluations_path = config['evaluations_path']
    output_path = config['output_path']
    column_mapping = config['column_mapping']
    target_mean = config.get('target_mean', 15.0)
    target_std = config.get('target_std', 2.0)
    professor_email_cfg = config.get('professor_email', 'jsboige@gmail.com')
    blacklisted_groups = config.get('blacklisted_groups', [])
    
    print("="*80)
    print(f"PIPELINE GRADEBOOK - {nom_classe}")
    print("="*80)
    print(f"\nFichier d'inscription : {os.path.basename(inscriptions_path)}")
    print(f"Fichier d'évaluation : {os.path.basename(evaluations_path)}")
    print(f"Fichier de sortie : {os.path.basename(output_path)}")
    print(f"Moyenne cible : {target_mean:.2f}")
    print(f"Écart-type cible : {target_std:.2f}")
    print()
    
    # Vérification des fichiers
    if not os.path.exists(inscriptions_path):
        print(f"❌ ERREUR : Fichier d'inscription introuvable : {inscriptions_path}")
        return None
    
    if not os.path.exists(evaluations_path):
        print(f"❌ ERREUR : Fichier d'évaluation introuvable : {evaluations_path}")
        return None
    
    # Chargement des données
    print("📚 Chargement des données...")
    student_records = load_student_records(inscriptions_path, column_mapping)
    print(f"✅ {len(student_records)} étudiants chargés")
    
    project_evals = load_grades_from_file(
        evaluations_path,
        PROFESSOR_FULL_NAME,
        professor_email_cfg,
        SIMILARITY_THRESHOLD
    )
    print(f"✅ {len(project_evals)} évaluations chargées")
    
    # Création de l'objet d'évaluation
    project_evaluation = ProjectEvaluation(professor_email_cfg, student_records)
    
    # Regroupement par projet
    df = pd.DataFrame([vars(e) for e in project_evals])
    grouped = df.groupby(df['groupe'].astype(str))
    
    group_evals = []
    for name, group_df in grouped:
        evals_indices = group_df.index
        evals = [e for i, e in enumerate(project_evals) if i in evals_indices]
        members = [s for s in student_records for s_sujet in s.sujets if fuzzy_match_group(name, s_sujet)]
        group_evals.append(GroupEvaluation(name, evals, members))
    
    project_evaluation.grouped_evaluations = group_evals
    print(f"✅ {len(group_evals)} groupes/projets identifiés")
    
    # FILTRAGE 1 : Liste noire (groupes doublons connus)
    if blacklisted_groups:
        print(f"\n🚫 Application de la liste noire ({len(blacklisted_groups)} groupes)...")
        non_blacklisted = []
        for group_eval in project_evaluation.grouped_evaluations:
            if group_eval.groupe in blacklisted_groups:
                print(f"❌ Groupe '{group_eval.groupe}' BLACKLISTÉ (doublon connu)")
                continue
            non_blacklisted.append(group_eval)
        
        project_evaluation.grouped_evaluations = non_blacklisted
        print(f"✅ {len(non_blacklisted)} groupes conservés après blacklist")
    
    # FILTRAGE 2 : Supprimer les groupes fantômes (sans membres inscrits)
    print("\n🔍 Suppression des groupes fantômes (sans membres inscrits)...")
    groups_with_members = []
    for group_eval in project_evaluation.grouped_evaluations:
        if not group_eval.group_members:
            print(f"⚠️  Groupe '{group_eval.groupe}' ignoré (aucun membre inscrit - doublon probable)")
            continue
        groups_with_members.append(group_eval)
    
    project_evaluation.grouped_evaluations = groups_with_members
    print(f"✅ {len(groups_with_members)} groupes avec membres inscrits")
    
    # Filtrage des évaluations
    print("\n🔧 Filtrage des évaluations (auto-évaluations, notes invalides, etc.)...")
    project_evaluation.filter_evaluations()
    
    # Validation finale des groupes (ceux avec évaluation professeur)
    validated_groups = []
    for group_eval in project_evaluation.grouped_evaluations:
        if not any(e.is_teacher for e in group_eval.evaluations):
            print(f"⚠️  Groupe '{group_eval.groupe}' ignoré (aucune évaluation professeur)")
            continue
        validated_groups.append(group_eval)
    
    project_evaluation.grouped_evaluations = validated_groups
    print(f"✅ {len(validated_groups)} groupes validés au final")
    
    # Application du redressement
    apply_rectification(project_evaluation, target_mean, target_std)
    
    # Génération du fichier Excel (avec filtrage NaN)
    generate_excel_workbook(project_evaluation, student_records, output_path)
    
    # Affichage des notes finales
    print("\n" + "="*80)
    print("NOTES FINALES PAR ÉTUDIANT")
    print("="*80)
    
    student_notes = []
    for group_eval in project_evaluation.grouped_evaluations:
        for member in group_eval.group_members:
            student_notes.append({
                'nom': member.nom,
                'prenom': member.prenom,
                'projet': group_eval.groupe,
                'note': group_eval.note_rectifiee
            })
    
    student_notes.sort(key=lambda x: (x['nom'], x['prenom']))
    
    for s in student_notes:
        print(f"  {s['nom']:20} {s['prenom']:15} | {s['projet'][:35]:35} | {s['note']:.2f}")
    
    print("\n" + "="*80)
    print("✅ PIPELINE TERMINÉ AVEC SUCCÈS")
    print("="*80)
    print(f"\n📊 Fichier Excel : {output_path}")
    
    # Statistiques finales
    final_grades = [g.note_rectifiee for g in project_evaluation.grouped_evaluations]
    print(f"\n📈 Statistiques finales :")
    print(f"   Moyenne : {np.mean(final_grades):.2f} (cible : {target_mean:.2f})")
    print(f"   Écart-type : {np.std(final_grades):.2f} (cible : {target_std:.2f})")
    print(f"   Min : {np.min(final_grades):.2f}")
    print(f"   Max : {np.max(final_grades):.2f}")
    
    return project_evaluation


def process_grades(registration_file, evaluation_file, output_dir, class_name, all_students):
    """
    Traite les notes pour une classe donnée à partir des fichiers d'inscription et d'évaluation (modèle EPITA).
    """
    # --- Chargement des données ---
    # `student_records` contient uniquement les étudiants de la classe en cours, pour former les groupes.
    student_records = load_student_records(registration_file)
    project_evals = load_grades_from_file(evaluation_file, PROFESSOR_FULL_NAME, professor_email, SIMILARITY_THRESHOLD)
    
    # --- Traitement ---
    # `project_evaluation` utilise `all_students` (toutes classes confondues) pour la validation des évaluateurs.
    project_evaluation = ProjectEvaluation(professor_email, all_students)
    
    df = pd.DataFrame([vars(e) for e in project_evals])
    # Group by the 'groupe' field, converting to string to be safe
    grouped = df.groupby(df['groupe'].astype(str))

    group_evals = []
    for name, group_df in grouped:
        evals_indices = group_df.index
        evals = [e for i, e in enumerate(project_evals) if i in evals_indices]
        
        # Utiliser la logique de correspondance floue pour trouver les membres
        members = [s for s in student_records for s_sujet in s.sujets if fuzzy_match_group(name, s_sujet)]
        group_evals.append(GroupEvaluation(name, evals, members))
    
    project_evaluation.grouped_evaluations = group_evals
    project_evaluation.filter_evaluations()

    # --- Points de contrôle post-filtrage ---
    validated_group_evals = []
    for group_eval in project_evaluation.grouped_evaluations:
        # Vérification 1: Le groupe ne doit pas être vide. Sinon, on l'ignore.
        if not group_eval.group_members:
            print(f"AVERTISSEMENT CRITIQUE: Le groupe '{group_eval.groupe}' n'a aucun membre inscrit et sera ignoré.")
            continue

        # Vérification 2: Présence de l'évaluation du professeur (maintenant bloquant)
        if not any(e.is_teacher for e in group_eval.evaluations):
            raise ValueError(f"Erreur critique : L'évaluation du professeur est manquante pour le groupe '{group_eval.groupe}'.")
        
        validated_group_evals.append(group_eval)
    project_evaluation.grouped_evaluations = validated_group_evals

    # --- Rectification des notes en deux passes ---
    target_mean, target_stdev = (15.5, 2.0)

    # Passe 1: Calcul des notes intermédiaires avec bonus/malus
    intermediate_grades = []
    for group_eval in project_evaluation.grouped_evaluations:
        raw_avg = group_eval.moyenne
        bonus_malus = palier_group_size_adjustment(len(group_eval.group_members))
        intermediate_grade = np.clip(raw_avg + bonus_malus, 0, 20)
        # Stockage temporaire de la note intermédiaire pour la passe 2
        group_eval.intermediate_grade = intermediate_grade
        intermediate_grades.append(intermediate_grade)

    # Passe 2: Normalisation basée sur les statistiques des notes intermédiaires
    if not intermediate_grades:
        # Gérer le cas où il n'y a aucune note à traiter
        project_mean = target_mean
        project_stdev = 0
    else:
        # Calcul des nouvelles statistiques basées sur les notes intermédiaires
        project_mean = np.mean(intermediate_grades)
        project_stdev = np.std(intermediate_grades)

    for group_eval in project_evaluation.grouped_evaluations:
        # Utilisation de la note intermédiaire stockée pour la normalisation
        current_grade = group_eval.intermediate_grade
        final_grade = adjust_grade(current_grade, project_mean, project_stdev, target_mean, target_stdev)
        group_eval.note_rectifiee = np.clip(final_grade, 14, 20)

    # --- Génération du fichier Excel ---
    output_filename = f"rapport_notes_{class_name}.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    # La fonction generate_workbook attend une liste de projets, nous l'encapsulons
    generate_workbook([project_evaluation], student_records, output_path)
    print(f"Le traitement pour la classe {class_name} est terminé. Fichier de résultats : {output_path.replace('//', '/')}")

    # Retourner l'objet traité pour l'analyse qualitative
    return project_evaluation


def generate_qualitative_summary(project_evaluation, class_name):
    """
    Affiche un résumé qualitatif des résultats de la notation pour une classe.
    """
    print(f"\n--- Résumé Qualitatif pour la classe : {class_name} ---")
    
    if not project_evaluation or not project_evaluation.grouped_evaluations:
        print("Aucune donnée d'évaluation à analyser.")
        return

    # Trier les groupes par note finale
    sorted_groups = sorted(project_evaluation.grouped_evaluations, key=lambda g: g.note_rectifiee, reverse=True)
    
    if not sorted_groups:
        print("Aucun groupe validé à analyser.")
        return

    # 1. Meilleur et moins bon groupe
    best_group = sorted_groups[0]
    worst_group = sorted_groups[-1]
    print(f"\n🏆 Meilleur Groupe : '{best_group.groupe}' (Note: {best_group.note_rectifiee:.2f})")
    print(f"📉 Moins bon Groupe : '{worst_group.groupe}' (Note: {worst_group.note_rectifiee:.2f})")

    # 2. Moyenne et Écart-type globaux
    overall_mean = project_evaluation.moyenne
    overall_std = project_evaluation.ecart_type
    print(f"\n📊 Statistiques Globales (basées sur les moyennes brutes de groupe pondérées) :")
    print(f"   - Moyenne de la classe : {overall_mean:.2f}")
    print(f"   - Écart-type de la classe : {overall_std:.2f}")

    # 3. Forme de la distribution
    final_notes = [g.note_rectifiee for g in sorted_groups]
    final_mean = np.mean(final_notes)
    final_median = np.median(final_notes)
    
    distribution_shape = "plutôt symétrique"
    if final_mean > final_median + 0.2:
        distribution_shape = "étirée vers la droite (quelques très bonnes notes)"
    elif final_mean < final_median - 0.2:
        distribution_shape = "étirée vers la gauche (quelques très mauvaises notes)"
        
    print(f"\n📈 Distribution des Notes Finales :")
    print(f"   - Moyenne : {final_mean:.2f}")
    print(f"   - Médiane : {final_median:.2f}")
    print(f"   - Forme : La distribution est {distribution_shape}.")

    # 4. Liste des sujets et résultats
    print("\n📋 Détail des notes par groupe (trié par note finale) :")
    print("-" * 60)
    print(f"{'Groupe (Sujet)':<40} {'Note Finale':>15}")
    print("-" * 60)
    for group in sorted_groups:
        print(f"{str(group.groupe):<40} {group.note_rectifiee:>15.2f}")
    print("-" * 60)


# =============================================================================
# SUPPORT MULTI-ÉPREUVES (Compatible avec le notebook GradeBook.ipynb)
# =============================================================================

def load_students_with_multiple_groups(file_path, epreuves_config):
    """
    Charge les étudiants depuis un fichier CSV avec support de plusieurs colonnes de groupes.
    
    Args:
        file_path: Chemin du fichier CSV d'inscription
        epreuves_config: Liste de dictionnaires définissant chaque épreuve
                        [{'nom': 'CC1', 'inscription_col': 'Groupe CC1', ...}, ...]
    
    Returns:
        DataFrame pandas avec les données des étudiants
    """
    df = pd.read_csv(file_path, delimiter=',')
    
    # Normaliser les noms de colonnes
    df.columns = df.columns.str.strip()
    
    # Mapping des colonnes standards
    col_mapping = {
        'Prénom': 'prenom',
        'Prenom': 'prenom',
        'Nom de famille': 'nom',
        'Nom': 'nom',
        'Adresse de courriel': 'email',
        'Adresse e-mail': 'email',
        'Email': 'email'
    }
    
    # Ajouter le mapping pour chaque épreuve
    for epreuve in epreuves_config:
        original_col = epreuve['inscription_col']
        target_col = f"groupe_{epreuve['nom'].lower().replace(' ', '_')}"
        if original_col in df.columns:
            col_mapping[original_col] = target_col
    
    df = df.rename(columns=col_mapping)
    
    return df


def load_grades_for_epreuve(file_path, epreuve_config, threshold=80):
    """
    Charge les évaluations pour une épreuve spécifique.
    
    Args:
        file_path: Chemin du fichier CSV d'évaluations
        epreuve_config: Configuration de l'épreuve
        threshold: Seuil de similarité pour identifier le professeur
    
    Returns:
        Liste d'objets EvaluationRecord
    """
    df = pd.read_csv(file_path, delimiter=',')
    
    # Mapping de colonnes standard (compatible Google Forms)
    column_mapping = {
        "Horodateur": "date",
        "Adresse e-mail": "email",
        "Votre nom": "nom",
        "Votre prénom": "prenom",
        "Groupe à évaluer": "groupe",
        "Sujet de la présentation": "sujet_libre",
        "Qualité de la présentation (communication, la forme)": "note_communication",
        "Qualité théorique (principes utilisés, classe d'algorithmes, contexte et explications des performances et des problèmes, histoire etc.)": "note_theorique",
        "Qualité technique (livrables, commits, qualité du code, démos, résultats, perspectives)": "note_technique",
        "Organisation (planning, répartition des tâches, collaboration, intégration au projet Github)": "note_organisation",
        "Points positifs de la présentation": "points_positifs",
        "Points négatifs de la présentation": "points_negatifs",
        "Recommandations pour s'améliorer": "recommandations"
    }
    
    # Gérer les colonnes tronquées (Google Forms parfois tronque les noms)
    for col in df.columns:
        for key in list(column_mapping.keys()):
            if col.startswith(key[:30]) or key.startswith(col[:30]):
                column_mapping[col] = column_mapping.get(key, col)
    
    df.rename(columns=column_mapping, inplace=True)
    
    # Créer le nom complet pour le matching
    df['nom_complet'] = df['prenom'].astype(str).str.strip() + " " + df['nom'].astype(str).str.strip()
    
    def normalize_name(name):
        return unidecode(name).lower().strip()
    
    normalized_prof_name = normalize_name(PROFESSOR_FULL_NAME)
    
    evaluations = []
    note_cols = [col for col in ["note_communication", "note_theorique", "note_technique", "note_organisation"] if col in df.columns]
    
    for _, row in df.iterrows():
        if not any(pd.notna(row.get(col)) and str(row.get(col)).strip() for col in note_cols):
            continue
        
        evaluator_name = row.get('nom_complet', '')
        if pd.isna(evaluator_name):
            evaluator_name = ''
        normalized_evaluator = normalize_name(str(evaluator_name))
        similarity = fuzz.ratio(normalized_prof_name, normalized_evaluator)
        
        is_teacher = similarity >= threshold
        
        notes_dict = {col: row.get(col) for col in note_cols}
        
        record = EvaluationRecord(
            row.get('date'), row.get('email'), row.get('nom'), row.get('prenom'),
            row.get('groupe'), row.get('sujet_libre'), notes_dict,
            row.get('points_positifs'), row.get('points_negatifs'), row.get('recommandations')
        )
        record.is_teacher = is_teacher
        evaluations.append(record)
    
    return evaluations


def process_epreuve(students_df, evaluations, epreuve_config, professor_email_cfg):
    """
    Traite une épreuve unique et retourne un ProjectEvaluation.
    
    Args:
        students_df: DataFrame des étudiants
        evaluations: Liste des EvaluationRecord
        epreuve_config: Configuration de l'épreuve
        professor_email_cfg: Email du professeur
    
    Returns:
        Objet ProjectEvaluation
    """
    epreuve_nom = epreuve_config['nom']
    groupe_col = f"groupe_{epreuve_nom.lower().replace(' ', '_')}"
    
    # Créer les StudentRecord pour cette épreuve
    student_records = []
    for _, row in students_df.iterrows():
        prenom = row.get('prenom')
        nom = row.get('nom')
        groupe = row.get(groupe_col)
        
        if pd.isna(prenom) or pd.isna(nom) or not str(prenom).strip() or not str(nom).strip():
            continue
        
        if pd.notna(groupe) and str(groupe).strip():
            student = StudentRecord(str(prenom), str(nom), str(groupe))
            student_records.append(student)
    
    # Créer l'objet ProjectEvaluation
    project_evaluation = ProjectEvaluation(professor_email_cfg, student_records)
    
    # Regrouper par groupe d'évaluation
    df_evals = pd.DataFrame([vars(e) for e in evaluations])
    if df_evals.empty:
        return project_evaluation
    
    grouped = df_evals.groupby(df_evals['groupe'].astype(str))
    
    group_evals = []
    for name, group_df in grouped:
        evals_indices = group_df.index
        evals = [e for i, e in enumerate(evaluations) if i in evals_indices]
        
        # Trouver les membres du groupe
        members = [s for s in student_records if fuzzy_match_group(name, s.sujets[0])]
        
        group_evals.append(GroupEvaluation(name, evals, members))
    
    project_evaluation.grouped_evaluations = group_evals
    
    # Filtrer les groupes sans membres
    project_evaluation.grouped_evaluations = [
        g for g in project_evaluation.grouped_evaluations if g.group_members
    ]
    
    # Filtrer les évaluations invalides
    project_evaluation.filter_evaluations()
    
    # Filtrer les groupes sans évaluation professeur
    project_evaluation.grouped_evaluations = [
        g for g in project_evaluation.grouped_evaluations
        if any(e.is_teacher for e in g.evaluations)
    ]
    
    return project_evaluation


def generate_multi_epreuve_excel(epreuves_results, students_df, epreuves_config, output_path):
    """
    Génère le fichier Excel avec plusieurs onglets :
    - Un onglet "Résumé Étudiants" avec toutes les notes + moyenne
    - Un onglet feedback par épreuve
    
    Args:
        epreuves_results: Liste de tuples (epreuve_config, ProjectEvaluation)
        students_df: DataFrame des étudiants
        epreuves_config: Liste des configurations d'épreuves
        output_path: Chemin du fichier Excel de sortie
    """
    
    print("\n" + "="*80)
    print("GÉNÉRATION DU FICHIER EXCEL MULTI-ÉPREUVES")
    print("="*80)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # === FEUILLE 1 : Résumé Étudiants ===
    summary_sheet = wb.create_sheet("Résumé Étudiants")
    
    # Construire les en-têtes dynamiquement
    headers = ["Nom", "Prénom"]
    for epreuve_cfg, _ in epreuves_results:
        headers.append(f"Groupe {epreuve_cfg['nom']}")
        headers.append(f"Note {epreuve_cfg['nom']}")
    headers.append("Moyenne Finale")
    
    for col, header in enumerate(headers, 1):
        summary_sheet.cell(1, col, header)
    
    # Style des en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(headers) + 1):
        cell = summary_sheet.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Collecter les notes par étudiant
    student_notes = {}
    for epreuve_cfg, proj_eval in epreuves_results:
        epreuve_nom = epreuve_cfg['nom']
        poids = epreuve_cfg.get('poids', 1.0)
        
        for group_eval in proj_eval.grouped_evaluations:
            for member in group_eval.group_members:
                key = (member.nom, member.prenom)
                if key not in student_notes:
                    student_notes[key] = {'nom': member.nom, 'prenom': member.prenom, 'notes': {}}
                student_notes[key]['notes'][epreuve_nom] = {
                    'groupe': group_eval.groupe,
                    'note': group_eval.note_rectifiee,
                    'poids': poids
                }
    
    # Trier les étudiants par nom et prénom
    sorted_students = sorted(student_notes.values(), key=lambda s: (s['nom'], s['prenom']))
    
    row = 2
    for student in sorted_students:
        summary_sheet.cell(row, 1, student['nom'])
        summary_sheet.cell(row, 2, student['prenom'])
        
        col = 3
        total_weighted = 0
        total_poids = 0
        
        for epreuve_cfg, _ in epreuves_results:
            epreuve_nom = epreuve_cfg['nom']
            note_info = student['notes'].get(epreuve_nom)
            
            if note_info:
                summary_sheet.cell(row, col, note_info['groupe'])
                summary_sheet.cell(row, col + 1, round(note_info['note'], 1))
                summary_sheet.cell(row, col + 1).number_format = '0.0'
                total_weighted += note_info['note'] * note_info['poids']
                total_poids += note_info['poids']
            else:
                summary_sheet.cell(row, col, "N/A")
                summary_sheet.cell(row, col + 1, "N/A")
            
            col += 2
        
        # Calculer la moyenne pondérée
        moyenne = total_weighted / total_poids if total_poids > 0 else 0
        summary_sheet.cell(row, col, round(moyenne, 1))
        summary_sheet.cell(row, col).number_format = '0.0'
        
        row += 1
    
    # Ajuster les largeurs de colonnes
    for i in range(1, len(headers) + 1):
        summary_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
    
    print(f"✅ Feuille 'Résumé Étudiants' créée ({row - 2} étudiants)")
    
    # === FEUILLES FEEDBACK (une par épreuve) ===
    for epreuve_cfg, proj_eval in epreuves_results:
        epreuve_nom = epreuve_cfg['nom']
        sheet_name = f"{epreuve_nom} Feedback"
        
        feedback_sheet = wb.create_sheet(sheet_name)
        
        # En-têtes
        feedback_sheet.cell(1, 1, "Groupe")
        feedback_sheet.cell(1, 2, "Points positifs")
        feedback_sheet.cell(1, 3, "Points négatifs")
        feedback_sheet.cell(1, 4, "Recommandations")
        
        # Style des en-têtes
        for col in range(1, 5):
            cell = feedback_sheet.cell(1, col)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Contenu
        feedback_row = 2
        feedback_count = 0
        
        for group_eval in proj_eval.grouped_evaluations:
            for evaluation in group_eval.evaluations:
                if is_feedback_empty(evaluation):
                    continue
                
                feedback_sheet.cell(feedback_row, 1, group_eval.groupe)
                
                # Nettoyer les valeurs NaN
                pp = getattr(evaluation, 'points_positifs', '')
                pn = getattr(evaluation, 'points_negatifs', '')
                rec = getattr(evaluation, 'recommandations', '')
                
                feedback_sheet.cell(feedback_row, 2, '' if pd.isna(pp) else str(pp))
                feedback_sheet.cell(feedback_row, 3, '' if pd.isna(pn) else str(pn))
                feedback_sheet.cell(feedback_row, 4, '' if pd.isna(rec) else str(rec))
                
                for col in range(1, 5):
                    cell = feedback_sheet.cell(feedback_row, col)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                feedback_row += 1
                feedback_count += 1
        
        # Ajuster les largeurs
        feedback_sheet.column_dimensions['A'].width = 35
        feedback_sheet.column_dimensions['B'].width = 40
        feedback_sheet.column_dimensions['C'].width = 40
        feedback_sheet.column_dimensions['D'].width = 40
        
        print(f"✅ Feuille '{sheet_name}' créée ({feedback_count} feedbacks)")
    
    # Sauvegarder
    wb.save(output_path)
    print(f"\n✅ Fichier Excel généré : {output_path}")


def run_multi_epreuve_pipeline(config):
    """
    Exécute le pipeline multi-épreuves complet.
    
    Args:
        config: Dictionnaire de configuration contenant :
            - nom_classe: Nom de la classe
            - inscriptions_path: Chemin du fichier CSV d'inscription
            - epreuves: Liste de configurations d'épreuves
            - output_path: Chemin du fichier Excel de sortie
            - professor_email: Email du professeur
    
    Returns:
        Liste de tuples (epreuve_config, ProjectEvaluation)
    """
    
    nom_classe = config['nom_classe']
    inscriptions_path = config['inscriptions_path']
    epreuves = config['epreuves']
    output_path = config['output_path']
    professor_email_cfg = config.get('professor_email', 'jsboige@gmail.com')
    
    print("="*80)
    print(f"PIPELINE GRADEBOOK MULTI-ÉPREUVES - {nom_classe}")
    print("="*80)
    print(f"\nFichier d'inscription : {os.path.basename(inscriptions_path)}")
    print(f"Nombre d'épreuves : {len(epreuves)}")
    for ep in epreuves:
        print(f"  - {ep['nom']} (poids: {ep.get('poids', 1.0)*100:.0f}%)")
    print(f"Fichier de sortie : {os.path.basename(output_path)}")
    print()
    
    # Vérification des fichiers
    if not os.path.exists(inscriptions_path):
        print(f"❌ ERREUR : Fichier d'inscription introuvable : {inscriptions_path}")
        return None
    
    for ep in epreuves:
        if not os.path.exists(ep['evaluations_path']):
            print(f"❌ ERREUR : Fichier d'évaluation introuvable pour {ep['nom']}: {ep['evaluations_path']}")
            return None
    
    # Chargement des étudiants
    print("📚 Chargement des inscriptions...")
    students_df = load_students_with_multiple_groups(inscriptions_path, epreuves)
    print(f"✅ {len(students_df)} étudiants chargés")
    
    # Traitement de chaque épreuve
    epreuves_results = []
    
    for epreuve_cfg in epreuves:
        print(f"\n{'='*60}")
        print(f"TRAITEMENT ÉPREUVE : {epreuve_cfg['nom']}")
        print(f"{'='*60}")
        
        # Charger les évaluations
        evaluations = load_grades_for_epreuve(
            epreuve_cfg['evaluations_path'],
            epreuve_cfg
        )
        print(f"✅ {len(evaluations)} évaluations chargées")
        
        # Traiter l'épreuve
        proj_eval = process_epreuve(students_df, evaluations, epreuve_cfg, professor_email_cfg)
        print(f"✅ {len(proj_eval.grouped_evaluations)} groupes validés")
        
        # Appliquer la rectification
        target_mean = epreuve_cfg.get('target_mean', 15.0)
        target_std = epreuve_cfg.get('target_std', 2.0)
        skip_normalization = epreuve_cfg.get('skip_normalization', False)

        if proj_eval.grouped_evaluations:
            apply_rectification(proj_eval, target_mean, target_std, skip_normalization)
        
        epreuves_results.append((epreuve_cfg, proj_eval))
    
    # Générer le fichier Excel multi-épreuves
    generate_multi_epreuve_excel(epreuves_results, students_df, epreuves, output_path)
    
    # Afficher le résumé
    print("\n" + "="*80)
    print("RÉSUMÉ DES NOTES")
    print("="*80)
    
    # Collecter toutes les notes
    all_student_notes = {}
    for epreuve_cfg, proj_eval in epreuves_results:
        epreuve_nom = epreuve_cfg['nom']
        poids = epreuve_cfg.get('poids', 1.0)
        
        for group_eval in proj_eval.grouped_evaluations:
            for member in group_eval.group_members:
                key = (member.nom, member.prenom)
                if key not in all_student_notes:
                    all_student_notes[key] = {'nom': member.nom, 'prenom': member.prenom, 'notes': {}}
                all_student_notes[key]['notes'][epreuve_nom] = {
                    'note': group_eval.note_rectifiee,
                    'poids': poids
                }
    
    # Afficher le tableau
    print(f"\n{'Nom':<20} {'Prénom':<15}", end="")
    for ep in epreuves:
        print(f" {ep['nom']:<12}", end="")
    print(f" {'Moyenne':<10}")
    print("-" * (50 + 12 * len(epreuves)))
    
    all_moyennes = []
    for key in sorted(all_student_notes.keys()):
        student = all_student_notes[key]
        print(f"{student['nom']:<20} {student['prenom']:<15}", end="")
        
        total_weighted = 0
        total_poids = 0
        
        for ep in epreuves:
            note_info = student['notes'].get(ep['nom'])
            if note_info:
                print(f" {note_info['note']:>10.2f}", end="")
                total_weighted += note_info['note'] * note_info['poids']
                total_poids += note_info['poids']
            else:
                print(f" {'N/A':>10}", end="")
        
        moyenne = total_weighted / total_poids if total_poids > 0 else 0
        all_moyennes.append(moyenne)
        print(f" {moyenne:>10.2f}")
    
    # Statistiques globales
    print("\n" + "="*80)
    print("STATISTIQUES GLOBALES")
    print("="*80)
    
    for epreuve_cfg, proj_eval in epreuves_results:
        notes = [g.note_rectifiee for g in proj_eval.grouped_evaluations]
        if notes:
            print(f"\n📊 {epreuve_cfg['nom']}:")
            print(f"   Moyenne: {np.mean(notes):.2f}")
            print(f"   Écart-type: {np.std(notes):.2f}")
            print(f"   Min: {np.min(notes):.2f} | Max: {np.max(notes):.2f}")
    
    if all_moyennes:
        print(f"\n📊 Moyenne générale de la classe:")
        print(f"   Moyenne: {np.mean(all_moyennes):.2f}")
        print(f"   Écart-type: {np.std(all_moyennes):.2f}")
        print(f"   Min: {np.min(all_moyennes):.2f} | Max: {np.max(all_moyennes):.2f}")
    
    print("\n" + "="*80)
    print("✅ PIPELINE MULTI-ÉPREUVES TERMINÉ AVEC SUCCÈS")
    print("="*80)
    
    return epreuves_results
